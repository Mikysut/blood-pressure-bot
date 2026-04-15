import re
import io
import logging
from html import escape
from datetime import datetime, timedelta

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    ContextTypes,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    filters,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import database
from database import MSK
import scheduler as sched
from charts import build_chart, PERIODS

logger = logging.getLogger(__name__)

# ── ConversationHandler состояния ────────────────────────────────────────────
MEASURE_BP, MEASURE_PULSE, MEASURE_NOTE, MEASURE_TIME, MEASURE_TIME_INPUT, MEASURE_CONFIRM = range(6)
REMIND_MORNING, REMIND_EVENING, REMIND_CONFIRM = range(10, 13)
EXCEL_WAIT_FILE, EXCEL_CONFIRM = range(20, 22)
CLEAR_CONFIRM = 30

# Кнопки главного меню
BTN_ADD      = "📝 Записать замер"
BTN_HISTORY  = "📋 История"
BTN_CHART    = "📈 График"
BTN_REMIND   = "⏰ Напоминания"
BTN_EXCEL    = "📥 Импорт Excel"
BTN_HELP     = "❓ Помощь"

BP_RE = re.compile(r"(\d{2,3})\s*/\s*(\d{2,3})")
TIME_RE = re.compile(r"^([01]\d|2[0-3]):([0-5]\d)$")
DATETIME_RE = re.compile(r"^(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2})$")


# ── Клавиатура ────────────────────────────────────────────────────────────────

def main_menu() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton(BTN_ADD)],
            [KeyboardButton(BTN_HISTORY), KeyboardButton(BTN_CHART)],
            [KeyboardButton(BTN_REMIND),  KeyboardButton(BTN_EXCEL)],
            [KeyboardButton(BTN_HELP)],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def skip_kb(text: str = "Пропустить") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton(text, callback_data="skip")]])


def confirm_kb(save_cb: str = "confirm_save") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Сохранить", callback_data=save_cb),
        InlineKeyboardButton("❌ Отмена",    callback_data="confirm_cancel"),
    ]])


def time_quick_kb(times: list[str], prefix: str) -> InlineKeyboardMarkup:
    buttons = [InlineKeyboardButton(t, callback_data=f"{prefix}_{t}") for t in times]
    return InlineKeyboardMarkup([buttons])


# ── /start ────────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    await database.register_user(user.id, user.username or user.first_name)
    sched.schedule_user(user.id, "09:00", "21:00")
    await update.message.reply_text(
        f"Привет, {escape(user.first_name)}!\n\n"
        "Я помогу отслеживать давление и пульс.\n"
        "Напоминания настроены на <b>09:00</b> и <b>21:00</b>.\n\n"
        "Используй кнопки меню ниже 👇",
        parse_mode="HTML",
        reply_markup=main_menu(),
    )


# ── Помощь ────────────────────────────────────────────────────────────────────

async def btn_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "<b>Как пользоваться ботом:</b>\n\n"
        "📝 <b>Записать замер</b> — пошаговый ввод давления и пульса\n"
        "📋 <b>История</b> — последние замеры\n"
        "📈 <b>График</b> — изменение давления за период\n"
        "⏰ <b>Напоминания</b> — настроить время напоминаний\n"
        "📥 <b>Импорт Excel</b> — загрузить старые данные из файла\n\n"
        "Используй кнопки меню для навигации.",
        parse_mode="HTML",
        reply_markup=main_menu(),
    )


# ── История — карусель ────────────────────────────────────────────────────────

def _carousel_text(r: dict, offset: int, total: int) -> str:
    dt        = datetime.fromisoformat(r["measured_at"]).strftime("%d.%m.%Y %H:%M")
    pulse_str = f"\nПульс: <b>{r['pulse']}</b> уд/мин" if r["pulse"] else ""
    note_str  = f"\n📝 {escape(r['note'])}" if r["note"] else ""
    pos       = f"<i>{offset + 1} из {total}</i>"
    return f"{pos}\n\n🕐 <b>{dt}</b>\n{r['systolic']}/{r['diastolic']} мм рт.ст.{pulse_str}{note_str}"


def _carousel_kb(r_id: int, offset: int, total: int) -> InlineKeyboardMarkup:
    nav_row = []
    if offset > 0:
        nav_row.append(InlineKeyboardButton("◀", callback_data=f"car_{offset - 1}"))
    nav_row.append(InlineKeyboardButton(f"{offset + 1} / {total}", callback_data="noop"))
    if offset + 1 < total:
        nav_row.append(InlineKeyboardButton("▶", callback_data=f"car_{offset + 1}"))
    rows = [nav_row]
    rows.append([InlineKeyboardButton("🗑 Удалить запись", callback_data=f"del_{r_id}_{offset}")])
    rows.append([InlineKeyboardButton("⚠️ Очистить всю историю", callback_data="clear_all")])
    return InlineKeyboardMarkup(rows)


async def btn_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    await database.register_user(user_id, update.effective_user.username or "")
    total = await database.get_total_count(user_id)
    if total == 0:
        await update.message.reply_text(
            "Записей пока нет.\nНажми 📝 Записать замер чтобы добавить первый замер.",
            reply_markup=main_menu(),
        )
        return
    records = await database.get_history(user_id, limit=1, offset=0)
    r = records[0]
    await update.message.reply_text(
        _carousel_text(r, 0, total),
        parse_mode="HTML",
        reply_markup=_carousel_kb(r["id"], 0, total),
    )


async def carousel_nav_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    offset = int(query.data.split("_")[1])
    user_id = update.effective_user.id
    total = await database.get_total_count(user_id)
    if total == 0:
        await query.message.edit_text("История пуста.", reply_markup=None)
        return
    offset = max(0, min(offset, total - 1))
    records = await database.get_history(user_id, limit=1, offset=offset)
    r = records[0]
    await query.message.edit_text(
        _carousel_text(r, offset, total),
        parse_mode="HTML",
        reply_markup=_carousel_kb(r["id"], offset, total),
    )


async def delete_record_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    parts = query.data.split("_")  # del_{id}_{offset}
    record_id = int(parts[1])
    offset    = int(parts[2])
    r = await database.get_measurement_by_id(record_id, update.effective_user.id)
    if not r:
        await query.message.edit_text("Запись не найдена.", reply_markup=None)
        return
    dt = datetime.fromisoformat(r["measured_at"]).strftime("%d.%m.%Y %H:%M")
    bp = f"{r['systolic']}/{r['diastolic']}"
    await query.message.edit_text(
        f"⚠️ Удалить эту запись?\n\n📅 <b>{dt}</b> — {bp} мм рт.ст.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Да, удалить", callback_data=f"delconfirm_{record_id}_{offset}"),
            InlineKeyboardButton("❌ Отмена",      callback_data=f"delcancel_{record_id}_{offset}"),
        ]]),
    )


async def delete_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    parts = query.data.split("_")  # delconfirm_{id}_{offset}
    record_id = int(parts[1])
    offset    = int(parts[2])
    user_id   = update.effective_user.id
    await database.delete_measurement(record_id, user_id)
    total = await database.get_total_count(user_id)
    if total == 0:
        await query.message.edit_text("✅ Запись удалена. История пуста.", reply_markup=None)
        return
    new_offset = min(offset, total - 1)
    records = await database.get_history(user_id, limit=1, offset=new_offset)
    r = records[0]
    await query.message.edit_text(
        _carousel_text(r, new_offset, total),
        parse_mode="HTML",
        reply_markup=_carousel_kb(r["id"], new_offset, total),
    )


async def delete_cancel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    parts = query.data.split("_")  # delcancel_{id}_{offset}
    record_id = int(parts[1])
    offset    = int(parts[2])
    user_id   = update.effective_user.id
    total = await database.get_total_count(user_id)
    r = await database.get_measurement_by_id(record_id, user_id)
    if not r:
        await query.message.edit_text("Запись не найдена.", reply_markup=None)
        return
    await query.message.edit_text(
        _carousel_text(r, offset, total),
        parse_mode="HTML",
        reply_markup=_carousel_kb(r["id"], offset, total),
    )


async def clear_all_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text(
        "Удалить <b>все</b> записи истории? Это действие нельзя отменить.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("⚠️ Да, удалить всё", callback_data="clear_all_confirm"),
            InlineKeyboardButton("❌ Отмена", callback_data="clear_all_cancel"),
        ]]),
    )


async def clear_all_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await database.clear_all_measurements(update.effective_user.id)
    await query.message.edit_text("✅ История очищена.", reply_markup=None)


async def clear_all_cancel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.delete()


# ── График ────────────────────────────────────────────────────────────────────

async def btn_chart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("7 дней",  callback_data="chart_week"),
        InlineKeyboardButton("30 дней", callback_data="chart_month"),
        InlineKeyboardButton("90 дней", callback_data="chart_3months"),
    ]])
    await update.message.reply_text("Выбери период:", reply_markup=keyboard)


async def chart_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    period = query.data.replace("chart_", "")
    user_id = update.effective_user.id

    days = PERIODS[period]
    since = datetime.now(tz=MSK).replace(tzinfo=None) - timedelta(days=days)
    records = await database.get_measurements_since(user_id, since)

    if len(records) < 2:
        await query.message.reply_text(
            "Недостаточно данных за выбранный период (нужно минимум 2 записи).",
            reply_markup=main_menu(),
        )
        return

    buf = build_chart(records, period)
    await query.message.reply_photo(photo=buf, caption="График давления")


# ── ConversationHandler: запись замера ────────────────────────────────────────

async def btn_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await database.register_user(update.effective_user.id, update.effective_user.username or "")
    context.user_data.clear()
    await update.message.reply_text(
        "Шаг 1/3 — Введи давление:\n\nФормат: <code>120/80</code>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Отмена", callback_data="measure_cancel")]]),
    )
    return MEASURE_BP


async def measure_bp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    match = BP_RE.search(text)
    if not match:
        await update.message.reply_text(
            "Не понял. Введи давление в формате <code>120/80</code>:",
            parse_mode="HTML",
        )
        return MEASURE_BP

    sys, dia = int(match.group(1)), int(match.group(2))
    if not (60 <= sys <= 250 and 40 <= dia <= 150):
        await update.message.reply_text("Значения вне диапазона. Систолическое: 60–250, диастолическое: 40–150.\nПовтори ввод:")
        return MEASURE_BP

    context.user_data["sys"] = sys
    context.user_data["dia"] = dia

    await update.message.reply_text(
        f"Давление: <b>{sys}/{dia}</b>\n\nШаг 2/3 — Введи пульс (уд/мин):",
        parse_mode="HTML",
        reply_markup=skip_kb("Пропустить пульс"),
    )
    return MEASURE_PULSE


async def measure_pulse_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    context.user_data["pulse"] = None
    await query.message.reply_text(
        "Шаг 3/3 — Добавь заметку (например: «после кофе»):",
        reply_markup=skip_kb("Без заметки"),
    )
    return MEASURE_NOTE


async def measure_pulse(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit() or not (30 <= int(text) <= 250):
        await update.message.reply_text(
            "Введи число от 30 до 250, или нажми «Пропустить»:",
            reply_markup=skip_kb("Пропустить пульс"),
        )
        return MEASURE_PULSE

    context.user_data["pulse"] = int(text)
    await update.message.reply_text(
        f"Пульс: <b>{text}</b>\n\nШаг 3/3 — Добавь заметку:",
        parse_mode="HTML",
        reply_markup=skip_kb("Без заметки"),
    )
    return MEASURE_NOTE


async def measure_note_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    context.user_data["note"] = None
    return await _ask_measure_time(query.message, context)


async def measure_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if len(text) > 500:
        await update.message.reply_text(
            "Заметка слишком длинная (максимум 500 символов). Сократи текст:",
            reply_markup=skip_kb("Без заметки"),
        )
        return MEASURE_NOTE
    context.user_data["note"] = text
    return await _ask_measure_time(update.message, context)


async def _ask_measure_time(message, context: ContextTypes.DEFAULT_TYPE):
    now = datetime.now(tz=MSK).replace(tzinfo=None)
    now_str = now.strftime("%d.%m.%Y %H:%M")
    await message.reply_text(
        f"Когда сделан замер?",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton(f"🕐 Сейчас ({now_str})", callback_data="measure_time_now"),
            InlineKeyboardButton("✏️ Ввести вручную",       callback_data="measure_time_manual"),
        ]]),
    )
    return MEASURE_TIME


async def measure_time_now(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    context.user_data["measured_at"] = None  # будет использовано now_msk() при сохранении
    return await _show_measure_confirm(query.message, context)


async def measure_time_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    await query.message.reply_text(
        "Введи дату и время замера:\n\nФормат: <code>ДД.ММ.ГГГГ ЧЧ:ММ</code>\nПример: <code>08.04.2026 09:30</code>",
        parse_mode="HTML",
    )
    return MEASURE_TIME_INPUT


async def measure_time_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    match = DATETIME_RE.match(text)
    if not match:
        await update.message.reply_text(
            "Неверный формат. Введи дату и время:\n<code>ДД.ММ.ГГГГ ЧЧ:ММ</code>",
            parse_mode="HTML",
        )
        return MEASURE_TIME_INPUT
    try:
        dt = datetime.strptime(text, "%d.%m.%Y %H:%M")
    except ValueError:
        await update.message.reply_text("Неверная дата. Попробуй ещё раз:")
        return MEASURE_TIME_INPUT
    context.user_data["measured_at"] = dt
    return await _show_measure_confirm(update.message, context)


async def _show_measure_confirm(message, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    pulse_str = f"\nПульс: <b>{d['pulse']}</b> уд/мин" if d.get("pulse") else ""
    note_str  = f"\nЗаметка: {escape(d['note'])}" if d.get("note") else ""
    if d.get("measured_at"):
        time_str = f"\nВремя: <b>{d['measured_at'].strftime('%d.%m.%Y %H:%M')}</b>"
    else:
        time_str = f"\nВремя: <b>сейчас</b>"
    await message.reply_text(
        f"Проверь данные:\n\n"
        f"Давление: <b>{d['sys']}/{d['dia']}</b> мм рт.ст.{pulse_str}{note_str}{time_str}\n\n"
        "Сохранить?",
        parse_mode="HTML",
        reply_markup=confirm_kb("measure_save"),
    )
    return MEASURE_CONFIRM


async def measure_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    d = context.user_data
    user_id = update.effective_user.id

    await database.add_measurement(user_id, d["sys"], d["dia"], d.get("pulse"), d.get("note"), d.get("measured_at"))

    eval_str = _evaluate_bp(d["sys"], d["dia"])
    await query.message.reply_text(
        f"✅ Записано!\n\n"
        f"Давление: <b>{d['sys']}/{d['dia']}</b> мм рт.ст."
        + (f"\nПульс: <b>{d['pulse']}</b> уд/мин" if d.get("pulse") else "")
        + (f"\nЗаметка: {escape(d['note'])}" if d.get("note") else "")
        + f"\n\n{eval_str}",
        parse_mode="HTML",
        reply_markup=main_menu(),
    )
    context.user_data.clear()
    return ConversationHandler.END


async def measure_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    context.user_data.clear()
    await query.message.reply_text("Отменено.", reply_markup=main_menu())
    return ConversationHandler.END


async def confirm_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    context.user_data.clear()
    await query.message.reply_text("Отменено.", reply_markup=main_menu())
    return ConversationHandler.END


# ── ConversationHandler: напоминания ─────────────────────────────────────────

async def btn_remind(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    await database.register_user(user_id, update.effective_user.username or "")
    user = await database.get_user(user_id)
    morning = user["remind_morning"] if user else "09:00"
    evening = user["remind_evening"] if user else "21:00"

    context.user_data["remind_morning"] = morning
    context.user_data["remind_evening"] = evening

    await update.message.reply_text(
        f"Текущие напоминания:\n"
        f"Утро: <b>{morning}</b>  |  Вечер: <b>{evening}</b>\n\n"
        "Выбери или введи новое <b>утреннее</b> время:",
        parse_mode="HTML",
        reply_markup=time_quick_kb(["07:00", "08:00", "09:00", "10:00"], "remind_m"),
    )
    return REMIND_MORNING


async def remind_morning_quick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    time_str = query.data.replace("remind_m_", "")
    context.user_data["remind_morning"] = time_str
    await query.message.reply_text(
        f"Утро: <b>{time_str}</b>\n\nВыбери или введи <b>вечернее</b> время:",
        parse_mode="HTML",
        reply_markup=time_quick_kb(["19:00", "20:00", "21:00", "22:00"], "remind_e"),
    )
    return REMIND_EVENING


async def remind_morning_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not TIME_RE.match(text):
        await update.message.reply_text(
            "Неверный формат. Введи время в формате ЧЧ:ММ (например: 08:30):",
            reply_markup=time_quick_kb(["07:00", "08:00", "09:00", "10:00"], "remind_m"),
        )
        return REMIND_MORNING
    context.user_data["remind_morning"] = text
    await update.message.reply_text(
        f"Утро: <b>{text}</b>\n\nВыбери или введи <b>вечернее</b> время:",
        parse_mode="HTML",
        reply_markup=time_quick_kb(["19:00", "20:00", "21:00", "22:00"], "remind_e"),
    )
    return REMIND_EVENING


async def remind_evening_quick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    time_str = query.data.replace("remind_e_", "")
    context.user_data["remind_evening"] = time_str
    return await _show_remind_confirm(query.message, context)


async def remind_evening_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not TIME_RE.match(text):
        await update.message.reply_text(
            "Неверный формат. Введи время в формате ЧЧ:ММ (например: 21:00):",
            reply_markup=time_quick_kb(["19:00", "20:00", "21:00", "22:00"], "remind_e"),
        )
        return REMIND_EVENING
    context.user_data["remind_evening"] = text
    return await _show_remind_confirm(update.message, context)


async def _show_remind_confirm(message, context: ContextTypes.DEFAULT_TYPE):
    m = context.user_data["remind_morning"]
    e = context.user_data["remind_evening"]
    await message.reply_text(
        f"Сохранить напоминания?\n\nУтро: <b>{m}</b>  |  Вечер: <b>{e}</b>",
        parse_mode="HTML",
        reply_markup=confirm_kb("remind_save"),
    )
    return REMIND_CONFIRM


async def remind_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    user_id = update.effective_user.id
    m = context.user_data["remind_morning"]
    e = context.user_data["remind_evening"]
    await database.update_remind_times(user_id, m, e)
    sched.schedule_user(user_id, m, e)
    await query.message.reply_text(
        f"✅ Напоминания сохранены:\nУтро: <b>{m}</b>  |  Вечер: <b>{e}</b>",
        parse_mode="HTML",
        reply_markup=main_menu(),
    )
    context.user_data.clear()
    return ConversationHandler.END


# ── ConversationHandler: импорт Excel ─────────────────────────────────────────

def _make_template() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Давление"

    headers = ["Дата", "Время", "Систолическое", "Диастолическое", "Пульс", "Заметка"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Пример строки
    ws.append(["01.04.2026", "09:15", 120, 80, 72, "после сна"])
    ws.append(["01.04.2026", "21:00", 125, 82, 74, ""])

    # Ширина колонок
    widths = [14, 10, 16, 18, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


async def btn_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📥 <b>Импорт из Excel</b>\n\nВыбери действие:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("📄 Скачать шаблон",  callback_data="excel_template"),
            InlineKeyboardButton("📤 Загрузить данные", callback_data="excel_upload"),
        ]]),
    )


async def excel_template(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    buf = _make_template()
    await query.message.reply_document(
        document=buf,
        filename="шаблон_давление.xlsx",
        caption=(
            "Заполни этот файл и загрузи обратно через кнопку «Загрузить данные».\n\n"
            "<b>Формат:</b>\n"
            "• Дата: ДД.ММ.ГГГГ\n"
            "• Время: ЧЧ:ММ\n"
            "• Пульс и Заметка — необязательны"
        ),
        parse_mode="HTML",
    )
    return ConversationHandler.END


async def excel_upload_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    await query.message.reply_text(
        "Отправь заполненный файл <b>шаблон_давление.xlsx</b>:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Отмена", callback_data="excel_cancel")]]),
    )
    return EXCEL_WAIT_FILE


async def excel_receive_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc or not doc.file_name.endswith((".xlsx", ".xls")):
        await update.message.reply_text("Нужен Excel-файл (.xlsx). Попробуй ещё раз:")
        return EXCEL_WAIT_FILE

    file = await doc.get_file()
    buf = io.BytesIO()
    await file.download_to_memory(buf)
    buf.seek(0)

    rows, errors = _parse_excel(buf)

    if not rows:
        await update.message.reply_text(
            f"Не удалось распознать данные в файле.\n"
            "Убедись что используешь шаблон от бота.\n"
            f"Ошибок: {errors}",
            reply_markup=main_menu(),
        )
        return ConversationHandler.END

    context.user_data["excel_rows"] = rows

    preview_lines = []
    for r in rows[:5]:
        dt = r["measured_at"].strftime("%d.%m.%Y %H:%M")
        p = f" пульс {r['pulse']}" if r.get("pulse") else ""
        preview_lines.append(f"• {dt} — {r['systolic']}/{r['diastolic']}{p}")

    preview = "\n".join(preview_lines)
    more = f"\n... и ещё {len(rows) - 5}" if len(rows) > 5 else ""
    err_str = f"\nПропущено строк с ошибками: {errors}" if errors else ""

    await update.message.reply_text(
        f"Найдено <b>{len(rows)}</b> записей:\n\n{preview}{more}{err_str}\n\nИмпортировать?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton(f"✅ Импортировать {len(rows)}", callback_data="excel_confirm"),
            InlineKeyboardButton("❌ Отмена", callback_data="excel_cancel"),
        ]]),
    )
    return EXCEL_CONFIRM


def _parse_excel(buf: io.BytesIO) -> tuple[list[dict], int]:
    wb = openpyxl.load_workbook(buf, data_only=True)
    ws = wb.active
    rows = []
    errors = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        try:
            date_val, time_val, sys_val, dia_val = row[0], row[1], row[2], row[3]
            pulse_val = row[4] if len(row) > 4 else None
            note_val  = str(row[5]).strip() if len(row) > 5 and row[5] else None

            # Парсим дату
            if isinstance(date_val, datetime):
                date_obj = date_val.date()
            else:
                date_obj = datetime.strptime(str(date_val).strip(), "%d.%m.%Y").date()

            # Парсим время
            if isinstance(time_val, datetime):
                time_obj = time_val.time()
            elif hasattr(time_val, "hour"):
                time_obj = time_val
            else:
                from datetime import time as dtime
                parts = str(time_val).strip().split(":")
                time_obj = dtime(int(parts[0]), int(parts[1]))

            measured_at = datetime.combine(date_obj, time_obj)

            systolic  = int(sys_val)
            diastolic = int(dia_val)
            pulse     = int(pulse_val) if pulse_val and str(pulse_val).strip() else None

            if not (60 <= systolic <= 250 and 40 <= diastolic <= 150):
                errors += 1
                continue

            rows.append({
                "measured_at": measured_at,
                "systolic":    systolic,
                "diastolic":   diastolic,
                "pulse":       pulse,
                "note":        note_val,
            })
        except (ValueError, KeyError, TypeError, IndexError, AttributeError) as e:
            logger.debug("Ошибка парсинга строки Excel: %s", e)
            errors += 1

    return rows, errors


async def excel_confirm_import(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    user_id = update.effective_user.id
    rows = context.user_data.pop("excel_rows", [])
    await database.add_measurements_bulk(user_id, rows)
    await query.message.reply_text(
        f"✅ Импортировано <b>{len(rows)}</b> записей!",
        parse_mode="HTML",
        reply_markup=main_menu(),
    )
    return ConversationHandler.END


async def excel_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.edit_reply_markup(reply_markup=None)
    context.user_data.pop("excel_rows", None)
    await query.message.reply_text("Отменено.", reply_markup=main_menu())
    return ConversationHandler.END


# ── Оценка давления ───────────────────────────────────────────────────────────

def _evaluate_bp(sys: int, dia: int) -> str:
    if sys < 90 or dia < 60:
        return "Давление низкое (гипотония)."
    elif sys <= 120 and dia <= 80:
        return "Отлично! Давление в норме."
    elif sys <= 129 and dia < 80:
        return "Давление немного повышено."
    elif sys <= 139 or dia <= 89:
        return "Высокое нормальное давление."
    else:
        return "Давление повышено. Рекомендуется проконсультироваться с врачом."


# ── Fallback для устаревших кнопок ────────────────────────────────────────────

async def stale_button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("Кнопка устарела. Начни действие заново.", show_alert=False)


# ── Глобальная отмена и обработчик ошибок ────────────────────────────────────

async def global_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Отменено.", reply_markup=main_menu())
    return ConversationHandler.END


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.error("Необработанное исключение", exc_info=context.error)
    if isinstance(update, Update) and update.effective_message:
        await update.effective_message.reply_text(
            "Произошла ошибка. Попробуй снова или нажми /start.",
            reply_markup=main_menu(),
        )


# ── Регистрация хендлеров ─────────────────────────────────────────────────────

def register_handlers(app):
    # Запись замера
    measure_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_ADD)}$"), btn_add)],
        states={
            MEASURE_BP: [MessageHandler(filters.TEXT & ~filters.COMMAND, measure_bp)],
            MEASURE_PULSE: [
                CallbackQueryHandler(measure_pulse_skip, pattern="^skip$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, measure_pulse),
            ],
            MEASURE_NOTE: [
                CallbackQueryHandler(measure_note_skip, pattern="^skip$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, measure_note),
            ],
            MEASURE_TIME: [
                CallbackQueryHandler(measure_time_now,    pattern="^measure_time_now$"),
                CallbackQueryHandler(measure_time_manual, pattern="^measure_time_manual$"),
            ],
            MEASURE_TIME_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, measure_time_input),
            ],
            MEASURE_CONFIRM: [
                CallbackQueryHandler(measure_save,   pattern="^measure_save$"),
                CallbackQueryHandler(confirm_cancel, pattern="^confirm_cancel$"),
            ],
        },
        fallbacks=[
            CallbackQueryHandler(measure_cancel, pattern="^measure_cancel$"),
            CommandHandler("cancel", global_cancel),
            CommandHandler("start", cmd_start),
        ],
        per_message=False,
    )

    # Напоминания
    remind_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(BTN_REMIND)}$"), btn_remind)],
        states={
            REMIND_MORNING: [
                CallbackQueryHandler(remind_morning_quick, pattern="^remind_m_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, remind_morning_text),
            ],
            REMIND_EVENING: [
                CallbackQueryHandler(remind_evening_quick, pattern="^remind_e_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, remind_evening_text),
            ],
            REMIND_CONFIRM: [
                CallbackQueryHandler(remind_save,    pattern="^remind_save$"),
                CallbackQueryHandler(confirm_cancel, pattern="^confirm_cancel$"),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", global_cancel),
            CommandHandler("start", cmd_start),
        ],
        per_message=False,
    )

    # Импорт Excel
    excel_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(excel_upload_prompt, pattern="^excel_upload$")],
        states={
            EXCEL_WAIT_FILE: [
                MessageHandler(filters.Document.ALL, excel_receive_file),
                CallbackQueryHandler(excel_cancel, pattern="^excel_cancel$"),
            ],
            EXCEL_CONFIRM: [
                CallbackQueryHandler(excel_confirm_import, pattern="^excel_confirm$"),
                CallbackQueryHandler(excel_cancel,         pattern="^excel_cancel$"),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", global_cancel),
            CommandHandler("start", cmd_start),
        ],
        per_message=False,
    )

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(measure_conv)
    app.add_handler(remind_conv)
    app.add_handler(excel_conv)

    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_HISTORY)}$"), btn_history))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_CHART)}$"),   btn_chart))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_EXCEL)}$"),   btn_excel))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_HELP)}$"),    btn_help))

    app.add_handler(CallbackQueryHandler(chart_callback,             pattern=r"^chart_"))
    app.add_handler(CallbackQueryHandler(carousel_nav_callback,     pattern=r"^car_\d+$"))
    app.add_handler(CallbackQueryHandler(delete_record_callback,    pattern=r"^del_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(delete_confirm_callback,   pattern=r"^delconfirm_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(delete_cancel_callback,    pattern=r"^delcancel_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(clear_all_callback,        pattern="^clear_all$"))
    app.add_handler(CallbackQueryHandler(clear_all_confirm_callback, pattern="^clear_all_confirm$"))
    app.add_handler(CallbackQueryHandler(clear_all_cancel_callback,  pattern="^clear_all_cancel$"))
    app.add_handler(CallbackQueryHandler(excel_template,            pattern="^excel_template$"))
    app.add_error_handler(error_handler)
    app.add_handler(CallbackQueryHandler(stale_button_callback))
